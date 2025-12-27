# export_excel.py
from __future__ import annotations
from typing import List, Dict, Any, Tuple, Iterable
from collections import defaultdict
from datetime import datetime, date, time

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

# İsteğe bağlı: öğretim elemanı adını Courses tablosundan çekebilmek için
try:
    from db import get_connection
except Exception:
    get_connection = None  # DB yoksa öğretim elemanı boş geçilir


def _instructor_map(course_ids: Iterable[int]) -> Dict[int, str]:
    """CourseID -> InstructorName haritası (varsa)."""
    if not get_connection or not course_ids:
        return {}
    ids = list({int(x) for x in course_ids})
    conn = get_connection(); cur = conn.cursor()
    q = f"SELECT CourseID, ISNULL(InstructorName,'') FROM dbo.Courses WHERE CourseID IN ({','.join('?'*len(ids))})"
    cur.execute(q, ids)
    mp = {int(cid): (name or '') for cid, name in cur.fetchall()}
    conn.close()
    return mp


def _styles():
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(bold=True, size=14)
    colhdr_font = Font(bold=True, color="FFFFFF")

    title_fill = PatternFill("solid", fgColor="F4A742")  # turuncu başlık bandı
    hdr_fill   = PatternFill("solid", fgColor="2D3748")  # koyu gri sütun başlığı

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center")
    wrap   = Alignment(wrap_text=True, vertical="center")

    return {
        "border": border_all,
        "title_font": title_font,
        "colhdr_font": colhdr_font,
        "title_fill": title_fill,
        "hdr_fill": hdr_fill,
        "center": center,
        "left": left,
        "wrap": wrap,
    }


def _auto_width(ws: Worksheet, cols: int):
    for c in range(1, cols + 1):
        letter = get_column_letter(c)
        maxlen = 0
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > maxlen:
                maxlen = len(s)
        ws.column_dimensions[letter].width = min(maxlen + 2, 45)


def export_schedule_to_excel(
    schedule: List[Dict[str, Any]],
    path: str,
    *,
    department_name: str | None = None,
) -> None:
    """
    schedule: ExamProgramPage'den gelen liste.
      Zorunlu alanlar: Date (date), Start (time), End (time), CourseCode, CourseName, ClassroomName, ExamType
      İsteğe bağlı: CourseID (öğretim elemanı çekmek için)
    path: kaydedilecek xlsx yolu
    """
    if not schedule:
        raise ValueError("Yazılacak program satırı yok.")

    # Grupla: Tarih -> satırlar (saat artan)
    grouped: Dict[date, List[Dict[str, Any]]] = defaultdict(list)
    for r in schedule:
        grouped[r["Date"]].append(r)
    for d in grouped:
        grouped[d].sort(key=lambda x: (x["Start"], x.get("CourseCode","")))

    # Ortak exam_type / bölüm adı
    exam_type = (schedule[0].get("ExamType") or "Sınav").upper()
    dept = department_name or "BİLGİSAYAR MÜHENDİSLİĞİ BÖLÜMÜ"

    # Öğretim elemanı haritası (varsa CourseID'ler)
    cid_list = [int(r.get("CourseID", 0)) for r in schedule if r.get("CourseID")]
    inst_map = _instructor_map(cid_list)

    st = _styles()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sınav Programı"

    # 1) Üst başlık (A1:E1 merge)
    ws.merge_cells("A1:E1")
    ws["A1"].value = f"{dept} {exam_type} SINAV PROGRAMI"
    ws["A1"].font = st["title_font"]
    ws["A1"].alignment = st["center"]
    ws["A1"].fill = st["title_fill"]
    ws.row_dimensions[1].height = 28

    # 2) Sütun başlıkları
    headers = ["Tarih", "Sınav Saati", "Ders Adı", "Öğretim Elemanı", "Derslik"]
    ws.append(headers)
    for col, _h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.font = st["colhdr_font"]
        cell.alignment = st["center"]
        cell.fill = st["hdr_fill"]
        cell.border = st["border"]
    ws.row_dimensions[2].height = 20

    # 3) Veri yazımı (tarih bloklarını hatırlayalım ki merge yapabilelim)
    start_row = 3
    blocks: List[Tuple[int, int]] = []  # (baş, son) satır aralıkları her tarih için

    for d in sorted(grouped.keys()):
        block_start = ws.max_row + 1 if ws.max_row >= start_row else start_row
        rows = grouped[d]

        for r in rows:
            ders_adi = f"{r.get('CourseName','')} ({r.get('CourseCode','')})"
            ogretim  = inst_map.get(int(r.get("CourseID", 0)), "")
            saat     = f"{r['Start'].strftime('%H:%M')} - {r['End'].strftime('%H:%M')}"
            derslik  = r.get("ClassroomName","")

            ws.append([
                d.strftime("%d.%m.%Y"),
                saat,
                ders_adi,
                ogretim,
                derslik
            ])

            # stil uygula
            last_row = ws.max_row
            for c in range(1, 6):
                cell = ws.cell(row=last_row, column=c)
                cell.border = st["border"]
                if c in (3, 4, 5):
                    cell.alignment = st["wrap"]
                elif c == 2:
                    cell.alignment = st["center"]
                else:
                    cell.alignment = st["left"]

        block_end = ws.max_row
        blocks.append((block_start, block_end))

    # 4) Tarih sütununu (A kolonu) blok bazlı birleştir (örnekteki gibi)
    for bstart, bend in blocks:
        if bstart < bend:
            ws.merge_cells(start_row=bstart, start_column=1, end_row=bend, end_column=1)
            cell = ws.cell(row=bstart, column=1)
            cell.alignment = st["center"]

    # 5) Sütun genişlikleri
    ws.column_dimensions["A"].width = 12   # Tarih
    ws.column_dimensions["B"].width = 16   # Saat
    ws.column_dimensions["C"].width = 44   # Ders Adı
    ws.column_dimensions["D"].width = 28   # Öğretim Elemanı
    ws.column_dimensions["E"].width = 26   # Derslik

    # Alternatif: otomatik genişlik (üstte sabit genişlikler daha kontrollü duruyor)
    # _auto_width(ws, 5)

    # 6) Kenarlık ve yükseklik ayarları
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 22

    # Kaydet
    wb.save(path)
